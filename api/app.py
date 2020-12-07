import openpyxl
from openpyxl import load_workbook
from flask  import Flask, render_template, request, session
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.sql import func
import datetime 
from datetime import date, timedelta
from flask_cors import CORS, cross_origin
import math, json
from werkzeug.datastructures import ImmutableMultiDict
from werkzeug.utils import secure_filename
from flask import jsonify
import simplejson as json
import time
import atexit
from apscheduler.schedulers.background import BackgroundScheduler
import glob
import os



app=Flask(__name__)
CORS(app)

app.config['SQLALCHEMY_DATABASE_URI']='postgresql://postgres:sherlock@localhost/pulse_collector'
# app.config['SQLALCHEMY_DATABASE_URI']='postgresql://postgres:postgres@localhost/pulse'
app.config['CORS_HEADERS'] = 'Content-Type'
db=SQLAlchemy(app)

# files_path="."
quarter_to_show= "Q12021"
current_quarter = "Q12021"

last_pulse_update_time = ""
last_sfdc_update_time = ""
last_support_update_time = ""

if(db is None):
    print("DB not init")

class Data(db.Model):
    __tablename__="pulse_data"
    id = db.Column(db.Integer, primary_key=True)
    PC_Cluster_UUID = db.Column(db.String(120))
    Customer_Name = db.Column(db.String(120))
    Account_Theater = db.Column(db.String(120))
    PC_VERSION = db.Column(db.String(10))
    CALM_VERSION = db.Column(db.String(10))
    EPSILON_VERSION = db.Column(db.String(10))
    Last_Reported_Date = db.Column(db.Date)
    ACTIVE_BP = db.Column(db.Integer)
    RUNNING_APP = db.Column(db.Integer)
    PROVISIONING_APP = db.Column(db.Integer)
    ERROR_APP = db.Column(db.Integer)
    DELETED_APP = db.Column(db.Integer)
    TOTAL_MANAGED_VMS = db.Column(db.Integer)
    TOTAL_AHV_VMS = db.Column(db.Integer)
    TOTAL_AWS_VMS =	db.Column(db.Integer)
    TOTAL_VMWARE_VMS = db.Column(db.Integer)
    TOTAL_GCP_VMS = db.Column(db.Integer)
    TOTAL_AZURE_VMS = db.Column(db.Integer)
    TOTAL_EXISTING_VMS = db.Column(db.Integer)
    TOTAL_K8S_POD = db.Column(db.Integer)
    ACTIVE_AHV_VMS = db.Column(db.Integer)
    ACTIVE_AWS_VMS = db.Column(db.Integer)
    ACTIVE_VMWARE_VMS = db.Column(db.Integer)
    ACTIVE_GCP_VMS = db.Column(db.Integer)
    ACTIVE_AZURE_VMS = db.Column(db.Integer)
    ACTIVE_EXISTING_VMS = db.Column(db.Integer)
    ACTIVE_K8S_POD = db.Column(db.Integer)
    LICENSE_VMS_COUNTS = db.Column(db.Integer)
    LICENSE_UNIQUE_VMS_COUNT = db.Column(db.Integer)
    LICENSE_REQUIRED_PACKS = db.Column(db.Integer)      
    QUARTER = db.Column(db.String(10))
    ADOPTION = db.Column(db.Float) 
    PERCENT_VMs_INUSE = db.Column(db.Float)     
    AWS_ACCOUNT = db.Column(db.Integer)
    GCP_ACCOUNT = db.Column(db.Integer)
    AZURE_ACCOUNT = db.Column(db.Integer)
    VMWARE_ACCOUNT = db.Column(db.Integer)
    PUBLIC_ACCOUNT = db.Column(db.Integer)

    def __init__(self, PC_Cluster_UUID, Customer_Name, Account_Theater, PC_VERSION, CALM_VERSION, EPSILON_VERSION, Last_Reported_Date, ACTIVE_BP, RUNNING_APP, PROVISIONING_APP, ERROR_APP,
        DELETED_APP, TOTAL_MANAGED_VMS, TOTAL_AHV_VMS, TOTAL_AWS_VMS, TOTAL_VMWARE_VMS, TOTAL_GCP_VMS, TOTAL_AZURE_VMS, TOTAL_EXISTING_VMS, TOTAL_K8S_POD, ACTIVE_AHV_VMS,
        ACTIVE_AWS_VMS, ACTIVE_VMWARE_VMS, ACTIVE_GCP_VMS, ACTIVE_AZURE_VMS, ACTIVE_EXISTING_VMS, ACTIVE_K8S_POD, LICENSE_VMS_COUNTS, LICENSE_UNIQUE_VMS_COUNT, LICENSE_REQUIRED_PACKS, QUARTER, ADOPTION, PERCENT_VMs_INUSE,
        AWS_ACCOUNT, VMWARE_ACCOUNT, AZURE_ACCOUNT, GCP_ACCOUNT, PUBLIC_ACCOUNT):
        
        self.PC_Cluster_UUID = PC_Cluster_UUID
        self.Customer_Name = Customer_Name
        self.Account_Theater = Account_Theater
        self.PC_VERSION = PC_VERSION
        self.CALM_VERSION = CALM_VERSION
        self.EPSILON_VERSION = EPSILON_VERSION
        self.Last_Reported_Date = Last_Reported_Date
        self.ACTIVE_BP = ACTIVE_BP
        self.RUNNING_APP = RUNNING_APP
        self.PROVISIONING_APP = PROVISIONING_APP
        self.ERROR_APP = ERROR_APP
        self.DELETED_APP = DELETED_APP
        self.TOTAL_MANAGED_VMS = TOTAL_MANAGED_VMS
        self.TOTAL_AHV_VMS = TOTAL_AHV_VMS
        self.TOTAL_AWS_VMS = TOTAL_AWS_VMS
        self.TOTAL_VMWARE_VMS = TOTAL_VMWARE_VMS
        self.TOTAL_GCP_VMS = TOTAL_GCP_VMS
        self.TOTAL_AZURE_VMS = TOTAL_AZURE_VMS
        self.TOTAL_EXISTING_VMS = TOTAL_EXISTING_VMS
        self.TOTAL_K8S_POD = TOTAL_K8S_POD
        self.ACTIVE_AHV_VMS = ACTIVE_AHV_VMS
        self.ACTIVE_AWS_VMS = ACTIVE_AWS_VMS
        self.ACTIVE_VMWARE_VMS = ACTIVE_VMWARE_VMS
        self.ACTIVE_GCP_VMS = ACTIVE_GCP_VMS
        self.ACTIVE_AZURE_VMS = ACTIVE_AZURE_VMS
        self.ACTIVE_EXISTING_VMS = ACTIVE_EXISTING_VMS
        self.ACTIVE_K8S_POD = ACTIVE_K8S_POD
        self.LICENSE_VMS_COUNTS = LICENSE_VMS_COUNTS
        self.LICENSE_UNIQUE_VMS_COUNT = LICENSE_UNIQUE_VMS_COUNT
        self.LICENSE_REQUIRED_PACKS = LICENSE_REQUIRED_PACKS
        self.QUARTER = QUARTER
        self.ADOPTION = ADOPTION
        self.PERCENT_VMs_INUSE = PERCENT_VMs_INUSE
        self.AWS_ACCOUNT = AWS_ACCOUNT
        self.AZURE_ACCOUNT = AZURE_ACCOUNT
        self.VMWARE_ACCOUNT = VMWARE_ACCOUNT
        self.GCP_ACCOUNT = GCP_ACCOUNT
        self.PUBLIC_ACCOUNT = PUBLIC_ACCOUNT

class SalesData(db.Model):
    __tablename__="sfdc_data"
    id = db.Column(db.Integer, primary_key=True)
    CUSTOMER_NAME = db.Column(db.String(120))
    TERM = db.Column(db.Integer())
    QTR_SOLD = db.Column(db.String(10))
    QTY_SOLD = db.Column(db.Integer())
    PRODUCT_CODE = db.Column(db.String(50))
    CALM_TCV = db.Column(db.String(50))

    def __init__(self, CUSTOMER_NAME, QTR_SOLD, PRODUCT_CODE, QTY_SOLD, CALM_TCV, TERM):
        
        self.CUSTOMER_NAME = CUSTOMER_NAME
        self.QTR_SOLD = QTR_SOLD
        self.PRODUCT_CODE = PRODUCT_CODE
        self.TERM = TERM
        self.QTY_SOLD = QTY_SOLD
        self.CALM_TCV = CALM_TCV

class SupportData(db.Model):
    __tablename__="support_data"
    id = db.Column(db.Integer, primary_key=True)
    CUSTOMER_NAME = db.Column(db.String(120))
    CASE_NUM = db.Column(db.Integer)
    DATE = db.Column(db.Date)

    def __init__(self, CUSTOMER_NAME, CASE_NUM, DATE):
        
        self.CUSTOMER_NAME = CUSTOMER_NAME
        self.CASE_NUM = CASE_NUM
        self.DATE = DATE
        
def quarter(i):
    switcher={
        1:'Q2',
        2:'Q3',
        3:'Q3',
        4:'Q3',
        5:'Q4',
        6:'Q4',
        7:'Q4',
        8:'Q1',
        9:'Q1',
        10: 'Q1',
        11: 'Q2',
        12: 'Q2'
    }
    return switcher.get(i)

def process_pulse(pulse_excel):
    wb = openpyxl.load_workbook(pulse_excel)
    print(wb.sheetnames)

    sheet = wb.active
    row_count = sheet.max_row
    column_count = sheet.max_column
    
    #getQuarter from filename
    print("Quarter is : ", pulse_excel[-17:-11])
    QUARTER = pulse_excel[-17:-11]

    version_dict = {}

    for row in sheet.iter_rows(min_row=2, min_col=1, max_row=row_count, max_col=column_count):  
        #print("PC_Cluster_UUID : ", row[0].value) #1
        PC_Cluster_UUID = row[0].value
        
        #print("Customer_Name : ", row[2].value)
        Customer_Name = row[2].value
        
        #print("Account_Theater : ", row[5].value) #6
        Account_Theater = row[5].value

        #print("PC_VERSION : ", row[8].value)    #9
        PC_VERSION = row[8].value

        #print("CALM_VERSION : ", row[9].value)   #10
        CALM_VERSION = row[9].value
        
        #print("EPSILON_VERSION : ", row[10].value) #11
        EPSILON_VERSION = row[10].value

        #print("Last_Reported_Date : ", row[11].value) #12
        Last_Reported_Date = row[11].value
        if(Last_Reported_Date is not None):
            date = datetime.datetime.strptime(Last_Reported_Date, '%d-%m-%Y')
        else:
            date = ""
        
        #print("ACTIVE_BP : ", row[12].value)       #13  
        ACTIVE_BP = row[12].value

        #print("RUNNING_APP : ", row[16].value)	    #16
        RUNNING_APP = row[16].value

        #print("PROVISIONING_APP : ", row[17].value) #18
        PROVISIONING_APP = row[17].value

        #print("ERROR_APP : ", row[18].value) #19	
        ERROR_APP = row[18].value

        #print("DELETED_APP : ", row[19].value)	#20
        DELETED_APP = row[19].value

        #print("TOTAL_MANAGED_VMS : ", row[20].value) #21
        TOTAL_MANAGED_VMS = row[20].value

        #print("TOTAL_AHV_VMS : ", row[21].value)	#22
        TOTAL_AHV_VMS = row[21].value

        #print("TOTAL_AWS_VMS : ", row[22].value) #23
        TOTAL_AWS_VMS = row[22].value
        
        #print("TOTAL_VMWARE_VMS : ", row[23].value) #24
        TOTAL_VMWARE_VMS = row[23].value
        
        #print("TOTAL_GCP_VMS : ", row[24].value) #25
        TOTAL_GCP_VMS = row[24].value
        
        #print("TOTAL_AZURE_VMS : ", row[25].value) #26
        TOTAL_AZURE_VMS = row[25].value
        
        #print("TOTAL_EXISTING_VMS : ", row[26].value) #27
        TOTAL_EXISTING_VMS = row[26].value
        
        #print("TOTAL_K8S_POD : ", row[27].value) #28
        TOTAL_K8S_POD = row[27].value
        
        #print("ACTIVE_AHV_VMS : ", row[28].value) #29
        ACTIVE_AHV_VMS = row[28].value
        
        #print("ACTIVE_AWS_VMS : ", row[29].value) #30
        ACTIVE_AWS_VMS = row[29].value
        
        #print("ACTIVE_VMWARE_VMS : ", row[30].value) #31
        ACTIVE_VMWARE_VMS = row[30].value
        
        #print("ACTIVE_GCP_VMS : ", row[31].value) #32
        ACTIVE_GCP_VMS = row[31].value
        
        #print("ACTIVE_AZURE_VMS : ", row[32].value) #33
        ACTIVE_AZURE_VMS = row[32].value
        
        #print("ACTIVE_EXISTING_VMS : ", row[33].value) #34
        ACTIVE_EXISTING_VMS = row[33].value
        
        #print("ACTIVE_K8S_POD : ", row[34].value) #35
        ACTIVE_K8S_POD = row[34].value
        
        #print("AWS_ACCOUNT : ", row[35].value) #36
        AWS_ACCOUNT = row[35].value
        
        #print("VMWARE_ACCOUNT : ", row[36].value) #37
        VMWARE_ACCOUNT = row[36].value
        
        #print("GCP_ACCOUNT : ", row[37].value) #38
        GCP_ACCOUNT = row[37].value
        
        #print("AZURE_ACCOUNT : ", row[38].value) #39
        AZURE_ACCOUNT = row[38].value

        #print("LICENSE_VMS_COUNTS : ", row[44].value) #45
        LICENSE_VMS_COUNTS = row[44].value
        
        #print("LICENSE_UNIQUE_VMS_COUNT : ", row[45].value) #46
        LICENSE_UNIQUE_VMS_COUNT = row[45].value
        
        #print("LICENSE_REQUIRED_PACKS : ", row[46].value) #47                     
        LICENSE_REQUIRED_PACKS = row[46].value                 

        if(Customer_Name is not None):

            if(("Nutanix" in Customer_Name) or ("Ntnx" in Customer_Name) or ("POC" in Customer_Name)):
                # print("Handling Ntnx acccounts. So skipping")
                continue

            if(CALM_VERSION in version_dict):
                x = version_dict[CALM_VERSION]
                version_dict[CALM_VERSION] = x+1  
            else:
                version_dict[CALM_VERSION] = 1
            
            # #current quarter
            # current_time = datetime.datetime.utcnow()
            # QUARTER = str(quarter(current_time.month)) + "'" + str(current_time.year)

            #Adoption
            if(LICENSE_UNIQUE_VMS_COUNT is not None):
                if(LICENSE_UNIQUE_VMS_COUNT != 0):
                    # ADOPTION = round((LICENSE_UNIQUE_VMS_COUNT/(LICENSE_REQUIRED_PACKS*25))*100,2)
                    ADOPTION = 0
                    
                else:
                    ADOPTION = -1    
            else:
                ADOPTION = -1

            #Softdelete
            if(TOTAL_AHV_VMS is not None):
                if(TOTAL_AHV_VMS != 0):
                    PERCENT_VMs_INUSE = round((ACTIVE_AHV_VMS/TOTAL_AHV_VMS)*100,2)
                
                else:
                    PERCENT_VMs_INUSE = -1   
            else:
                PERCENT_VMs_INUSE = -1
            
            #Public Account
            if(AWS_ACCOUNT > 0 or AZURE_ACCOUNT > 0 or GCP_ACCOUNT > 0):
                PUBLIC_ACCOUNT = AWS_ACCOUNT + AZURE_ACCOUNT + GCP_ACCOUNT
            
            else:
                PUBLIC_ACCOUNT = -1

            #Populate the Data
            data=Data(PC_Cluster_UUID, Customer_Name.capitalize(), Account_Theater, PC_VERSION, CALM_VERSION, EPSILON_VERSION, date, ACTIVE_BP, RUNNING_APP, PROVISIONING_APP, ERROR_APP, DELETED_APP, TOTAL_MANAGED_VMS, TOTAL_AHV_VMS, TOTAL_AWS_VMS, TOTAL_VMWARE_VMS, TOTAL_GCP_VMS, TOTAL_AZURE_VMS, TOTAL_EXISTING_VMS, TOTAL_K8S_POD, ACTIVE_AHV_VMS, ACTIVE_AWS_VMS, ACTIVE_VMWARE_VMS, ACTIVE_GCP_VMS, ACTIVE_AZURE_VMS, ACTIVE_EXISTING_VMS, ACTIVE_K8S_POD, LICENSE_VMS_COUNTS, LICENSE_UNIQUE_VMS_COUNT, LICENSE_REQUIRED_PACKS, QUARTER, ADOPTION, PERCENT_VMs_INUSE,
            AWS_ACCOUNT, VMWARE_ACCOUNT, AZURE_ACCOUNT, GCP_ACCOUNT, PUBLIC_ACCOUNT)
            
            db.session.add(data)
            # print("Entered Data for customer : ", Customer_Name)    
    
    db.session.commit()

def process_sfdc(sfdc_excel):
    wb = openpyxl.load_workbook(sfdc_excel)
    print(wb.sheetnames)

    sheet = wb.active
    row_count = sheet.max_row
    column_count = sheet.max_column
    print("Sheet Name is :", sheet)
    
    # print("Quarter is : ", excel[-11:-5])
    # QUARTER = excel[-11:-5]

    for row in sheet.iter_rows(min_row=2, min_col=1, max_row=row_count, max_col=column_count):  
        ACCOUNT_NAME = row[0].value
        
        QUARTER_SOLD = row[1].value
        
        PRODUCT_CODE = row[2].value
    
        QTY_SOLD = row[3].value

        CALM_TCV = row[7].value

        TERM = row[8].value

        if(ACCOUNT_NAME is not None):

            #Populate the Data
            salesData = SalesData(ACCOUNT_NAME.capitalize(), QUARTER_SOLD, PRODUCT_CODE, QTY_SOLD, CALM_TCV, TERM)
            
            db.session.add(salesData)
            # print("Entered Data for Account : ", ACCOUNT_NAME)    
    
    db.session.commit()

def process_support(support_excel):
    wb = openpyxl.load_workbook(support_excel)
    print(wb.sheetnames)

    sheet = wb.active
    row_count = sheet.max_row
    column_count = sheet.max_column

    for row in sheet.iter_rows(min_row=2, min_col=1, max_row=row_count, max_col=column_count):  
        ACCOUNT_NAME = row[1].value
        
        CASE_NUM = row[0].value
        
        OPEN_DATE = row[6].value

        if(ACCOUNT_NAME is not None):

            if(("Nutanix" in ACCOUNT_NAME) or ("Ntnx" in ACCOUNT_NAME) or ("POC" in ACCOUNT_NAME)):
                # print("Handling Ntnx acccounts. So skipping")
                continue

            #Populate the Data
            if(OPEN_DATE is not None):
                if(type(OPEN_DATE) is str):
                    tstr = OPEN_DATE.split("/")
                    openDate = tstr[2]+"-"+tstr[0]+"-"+tstr[1]
                else:
                    tstr = OPEN_DATE.strftime('%m-%d-%Y').split("-")
                    openDate = tstr[2]+"-"+tstr[1]+"-"+tstr[0]
                    
                # print("tstr ", tstr)
                # openDate = tstr[2]+"-"+tstr[0]+"-"+tstr[1]
                # print("Customer ", ACCOUNT_NAME)
                # print("Open Date ", openDate)
                # opendate = datetime.datetime.strptime(OPEN_DATE.strftime('%m-%d-%Y'), '%Y-%m-%d')
            
            supportData = SupportData(ACCOUNT_NAME.capitalize(), CASE_NUM, openDate)
            
            db.session.add(supportData)
            #print("Entered Data for Account : ", ACCOUNT_NAME)    
    
    db.session.commit()


@app.route("/")
def __intialize():
    print("Initialized loop")
    excelList = ["Calm_Customer_List_Q12020_pulse.xlsx", "Calm_Customer_List_Q22020_pulse.xlsx", "Calm_Customer_List_Q32020_pulse.xlsx", "Calm_Customer_List_Q42020_pulse.xlsx", "Calm_Customer_List_Q12021_pulse.xlsx"]
    for pulse_excel in excelList:
        print("Initialized loop")
        process_pulse(pulse_excel)

    excelList = ["Calm_Licenses_Sold-FY'21Q1_sfdc.xlsx"]
    for sfdc_excel in excelList:
        process_sfdc(sfdc_excel)

    excelList = ["Calm_cases.xlsx"]
    for support_excel in excelList:
        print("Initialized cases")
        process_support(support_excel)
        
    return ""

@app.route("/getStats", methods=['GET']) 
def getStats():
    stats = {}
    if request.method=='GET':
        # print(" Current Quarter is : %s " % quarter_to_show)

        num_of_active_customers = db.session.query(Data.Customer_Name).distinct().filter(Data.QUARTER == quarter_to_show).count()
        # print(" Total Active Customers: %s " % num_of_active_customers)
        
        num_of_active_BPs = db.session.query(func.sum(Data.ACTIVE_BP)).filter(Data.QUARTER == quarter_to_show).scalar()
        # print(" Total Active BPs: %s " % num_of_active_BPs)

        num_of_running_APPs = db.session.query(func.sum(Data.RUNNING_APP)).filter(Data.QUARTER == quarter_to_show).scalar()
        # print(" Total Running APPs: %s " % num_of_running_APPs)
        
        num_of_provisioning_APPs = db.session.query(func.sum(Data.PROVISIONING_APP)).filter(Data.QUARTER == quarter_to_show).scalar()
        # print(" Total Provisioning APPs: %s " % num_of_provisioning_APPs)
        
        num_of_managed_VMs= db.session.query(func.sum(Data.TOTAL_MANAGED_VMS)).filter(Data.QUARTER == quarter_to_show).scalar()
        # print(" Total Managed VMs: %s " % num_of_managed_VMs)
        
        num_of_active_AHV_VMs = db.session.query(func.sum(Data.ACTIVE_AHV_VMS)).filter(Data.QUARTER == quarter_to_show).scalar()
        # print(" Total Active AHV VMs: %s " % num_of_active_AHV_VMs)
        
        num_of_active_AWS_VMs = db.session.query(func.sum(Data.ACTIVE_AWS_VMS)).filter(Data.QUARTER == quarter_to_show).scalar()
        # print(" Total Active AWS VMs: %s " % num_of_active_AWS_VMs)
        
        num_of_active_VMWare_VMs = db.session.query(func.sum(Data.ACTIVE_VMWARE_VMS)).filter(Data.QUARTER == quarter_to_show).scalar()
        # print(" Total Active VMware VMs: %s " % num_of_active_VMWare_VMs)
        
        num_of_active_GCP_VMs = db.session.query(func.sum(Data.ACTIVE_GCP_VMS)).filter(Data.QUARTER == quarter_to_show).scalar()
        # print(" Total Active GCP VMs: %s " % num_of_active_GCP_VMs)
        
        num_of_active_Existing_VMs = db.session.query(func.sum(Data.ACTIVE_EXISTING_VMS)).filter(Data.QUARTER == quarter_to_show).scalar()
        # print(" Total Active Existing VMs: %s " % num_of_active_Existing_VMs)
        
        num_of_licensed_unique_VMs = db.session.query(func.sum(Data.LICENSE_UNIQUE_VMS_COUNT)).filter(Data.QUARTER == quarter_to_show).scalar()
        # print(" Total licenses unique VMs: %s " % num_of_licensed_unique_VMs)

        num_of_licenses_required = db.session.query(func.sum(Data.LICENSE_REQUIRED_PACKS)).filter(Data.QUARTER == quarter_to_show).scalar()
        # print(" Total licenses required: %s " % num_of_licenses_required)

        # average_adoption = db.session.query(func.avg(Data.ADOPTION)).filter(Data.ADOPTION > 0).filter(Data.QUARTER == quarter_to_show).scalar()

        # print(average_adoption)
        # print(type(average_adoption))
        # print(round(average_adoption, 2))

        unique_paid_customers = db.session.query(SalesData.CUSTOMER_NAME).distinct().count()
        
        total_licenses_sold = 0
        total_cores_sold = 0

        licenses_purchased = db.session.query(SalesData.PRODUCT_CODE, SalesData.QTY_SOLD).all()
        for record in licenses_purchased:
            if("CORE" in record.PRODUCT_CODE):
                total_cores_sold = total_cores_sold + record.QTY_SOLD
            else:
                total_licenses_sold = total_licenses_sold + record.QTY_SOLD

        # total_licenses_sold = db.session.query(func.sum(SalesData.QTY_SOLD)).scalar()

        total_tcv = db.session.query(func.sum(SalesData.CALM_TCV)).scalar()
        avg_term = db.session.query(func.avg(SalesData.TERM)).scalar()
        print("avg term ", round(avg_term, 2))
        stats['active_customers'] = num_of_active_customers
        stats['active_BPs'] = num_of_active_BPs
        stats['running_APPs'] = num_of_running_APPs
        stats['provisioning_APPs'] = num_of_provisioning_APPs
        stats['managed_VM'] = num_of_managed_VMs
        stats['active_AHV_VMs'] = num_of_active_AHV_VMs
        stats['active_AWS_VMs'] = num_of_active_AWS_VMs
        stats['active_VMWare_VMs'] = num_of_active_VMWare_VMs
        stats['active_GCP_VMs'] = num_of_active_GCP_VMs
        stats['active_Existing_VMs'] = num_of_active_Existing_VMs
        stats['licensed_unique_VMs'] = num_of_licensed_unique_VMs 
        stats['licenses_required'] = num_of_licenses_required
        # stats['avg_adoption'] = round(average_adoption, 2)
        stats['paid_customers'] = unique_paid_customers
        stats['avg_term'] = round(avg_term/12, 2)
        stats['lifetime_tcv'] = total_tcv
        stats['licenses_sold'] = total_licenses_sold
        stats['cores_sold'] = total_cores_sold

        #All Pulse Customers
        pulse_customers = db.session.query(Data.Customer_Name).distinct().filter(Data.QUARTER == quarter_to_show)
        #All Paid Customers
        sfdc_paid_customers = db.session.query(SalesData.CUSTOMER_NAME).distinct()

        unique_pulse_paid = 0
        for customer in sfdc_paid_customers:
            if(customer in pulse_customers):
                unique_pulse_paid += 1

        stats['pulse_paid'] = unique_pulse_paid

    return stats

@app.route("/getAdoption", methods=['GET'])
def getAdoption():
    #Prepare adoption Data
    adoption_records = []
    
    #Paid Customers & Licenses count    
    licenses_purchased = db.session.query(SalesData.CUSTOMER_NAME, func.sum(SalesData.QTY_SOLD).label("QTY_SOLD")).group_by(SalesData.CUSTOMER_NAME).all()
    # print(len(licenses_purchased))
    
    paid_customers = {}
    for record in licenses_purchased:
        if(record.CUSTOMER_NAME not in paid_customers):
            paid_customers[record.CUSTOMER_NAME] = record.QTY_SOLD
    
    # print(len(paid_customers))
    
    #Get records for adoption computation

    # Print date
    base_date = datetime.datetime.strptime("2020-05-01", '%Y-%m-%d')
    # print("Base date", base_date)
    # within_Range = db.session.query(Data.Customer_Name, Data.Last_Reported_Date).filter(Data.QUARTER == quarter_to_show).filter(Data.Last_Reported_Date > base_date)
    
    records = db.session.query(func.sum(Data.LICENSE_UNIQUE_VMS_COUNT).label("VMs"), Data.Customer_Name).filter(Data.LICENSE_UNIQUE_VMS_COUNT > 0).filter(Data.QUARTER == quarter_to_show).filter(Data.Last_Reported_Date > base_date).group_by(Data.Customer_Name).all()
    # records = db.session.query(func.sum(Data.LICENSE_UNIQUE_VMS_COUNT).label("VMs"), Data.Customer_Name).filter(Data.LICENSE_UNIQUE_VMS_COUNT > 0).filter(Data.QUARTER == quarter_to_show).group_by(Data.Customer_Name).all()
    # print(len(records))
    
    adoption_dict = {}
    for record in records:
        customer = record.Customer_Name
        if(paid_customers.get(customer) is not None):
            adoption_dict[customer] = round((record.VMs/((paid_customers.get(customer)+1)*25))*100,2)
            # print("Customer : " , customer + "VMs is : ", str(record.VMs) + "Licenses is : ", str(paid_customers.get(customer)) + " adoption is : ", adoption_dict[customer])
            db.session.query(Data).filter(Data.Customer_Name == customer).filter(Data.QUARTER == quarter_to_show).update({Data.ADOPTION : adoption_dict[customer]})
            db.session.commit()

    # print(len(adoption_dict))
    # sorted(adoption_dict.items(), key=lambda x: x[1], reverse=True)    
    # # print(adoption_dict)

    for key in adoption_dict:
        adoption_record = {}
        adoption_record["Name"] = key
        adoption_record["Value"] = adoption_dict[key]
        adoption_record["QTY"] = paid_customers.get(key)
        
        adoption_records.append(adoption_record)

    # print(adoption_records)    
    
    return jsonify({'adoption records': adoption_records})

@app.route("/getAverageAdoption", methods=['GET'])
def getAverageAdoption():
    average_adoption = db.session.query(func.avg(Data.ADOPTION)).filter(Data.ADOPTION > 0).filter(Data.QUARTER == quarter_to_show).scalar()
    
    stats = {}
    stats['avg_adoption'] = round(average_adoption, 2)

    return stats
    
@app.route("/getPaidCustomersList", methods=['GET'])
def getPaidCustomersList():

    licenses_purchased = db.session.query(SalesData.CUSTOMER_NAME).distinct().all()
    paid_customers = []
    for record in licenses_purchased:
        # print("Cust name ", record.CUSTOMER_NAME)
        if(record.CUSTOMER_NAME not in paid_customers):
            paid_customers.append(record.CUSTOMER_NAME)

    records = db.session.query(Data.Customer_Name).distinct().filter(Data.QUARTER == quarter_to_show).order_by(Data.Customer_Name.desc()).all()
    
    cust_list = []
    for record in records:
        customer = record.Customer_Name
        if(customer in paid_customers):
            cust_list.append(customer.capitalize())
        
    # print(cust_list)

    return jsonify({'Customer list': cust_list})

@app.route("/getCalmVersionDistro", methods=['GET'])
def getCalmVersionDistro():
    #Prepare Calm Version records
    version_records = []
    calm_versions_records = db.session.query(Data.Customer_Name, Data.CALM_VERSION).filter(Data.QUARTER == quarter_to_show).all()
    version_dict = {}

    for record in calm_versions_records:
        # print(record.__dict__)
        version = record.CALM_VERSION
        
        if(version in version_dict):
            num = version_dict[version]
            version_dict[version] = num+1
        elif(version is not None):
            version_dict[version] = 1

    # print(version_dict)

    for key in version_dict:
        version_record = {}
        version_record["Version_Name"] = key
        version_record["Value"] = version_dict[key]
        
        version_records.append(version_record)

    # print(version_records)

    return jsonify({'version records': version_records})

@app.route("/getStatsByQtr", methods=['GET']) 
def getStatsByQtr():
    statsByQtr = []
    if request.method=='GET':

        num_of_active_bps_q1 = db.session.query(func.sum(Data.ACTIVE_BP)).filter(Data.QUARTER == "Q12020").scalar()
        # # print(" Total Active BPs in Q1: %s " % num_of_active_bps_q1)

        num_of_active_bps_q2 = db.session.query(func.sum(Data.ACTIVE_BP)).filter(Data.QUARTER == "Q22020").scalar()
        # print(" Total Active BPs in Q2: %s " % num_of_active_bps_q2)

        num_of_active_bps_q3 = db.session.query(func.sum(Data.ACTIVE_BP)).filter(Data.QUARTER == "Q32020").scalar()
        # print(" Total Active BPs in Q3: %s " % num_of_active_bps_q3)

        num_of_active_bps_q4 = db.session.query(func.sum(Data.ACTIVE_BP)).filter(Data.QUARTER == "Q42020").scalar()
        # print(" Total Active BPs in Q4: %s " % num_of_active_bps_q4)

        num_of_active_bps_q5 = db.session.query(func.sum(Data.ACTIVE_BP)).filter(Data.QUARTER == "Q12021").scalar()
        # print(" Total Active BPs in Q5: %s " % num_of_active_bps_q4)

        num_of_running_APPs_q1 = db.session.query(func.sum(Data.RUNNING_APP)).filter(Data.QUARTER == "Q12020").scalar()
        # print(" Total Running APPs in Q1: %s " % num_of_running_APPs_q1)

        num_of_running_APPs_q2 = db.session.query(func.sum(Data.RUNNING_APP)).filter(Data.QUARTER == "Q22020").scalar()
        # print(" Total Running APPs in Q2: %s " % num_of_running_APPs_q2)

        num_of_running_APPs_q3 = db.session.query(func.sum(Data.RUNNING_APP)).filter(Data.QUARTER == "Q32020").scalar()
        # print(" Total Running APPs in Q3: %s " % num_of_running_APPs_q2)

        num_of_running_APPs_q4 = db.session.query(func.sum(Data.RUNNING_APP)).filter(Data.QUARTER == "Q42020").scalar()
        # print(" Total Running APPs in Q4: %s " % num_of_running_APPs_q2)

        num_of_running_APPs_q5 = db.session.query(func.sum(Data.RUNNING_APP)).filter(Data.QUARTER == "Q12021").scalar()
        # print(" Total Running APPs in Q5: %s " % num_of_running_APPs_q5)

        num_of_active_AHV_VMs_q1 = db.session.query(func.sum(Data.ACTIVE_AHV_VMS)).filter(Data.QUARTER == "Q12020").scalar()
        # print(" Total Active AHV VMs: %s " % num_of_active_AHV_VMs_q1)

        num_of_active_AHV_VMs_q2 = db.session.query(func.sum(Data.ACTIVE_AHV_VMS)).filter(Data.QUARTER == "Q22020").scalar()
        # print(" Total Active AHV VMs: %s " % num_of_active_AHV_VMs_q2)

        num_of_active_AHV_VMs_q3 = db.session.query(func.sum(Data.ACTIVE_AHV_VMS)).filter(Data.QUARTER == "Q32020").scalar()
        # print(" Total Active AHV VMs: %s " % num_of_active_AHV_VMs_q3)

        num_of_active_AHV_VMs_q4 = db.session.query(func.sum(Data.ACTIVE_AHV_VMS)).filter(Data.QUARTER == "Q42020").scalar()
        # print(" Total Active AHV VMs: %s " % num_of_active_AHV_VMs_q4)

        num_of_active_AHV_VMs_q5 = db.session.query(func.sum(Data.ACTIVE_AHV_VMS)).filter(Data.QUARTER == "Q12021").scalar()
        # print(" Total Active AHV VMs: %s " % num_of_active_AHV_VMs_q5)

        average_adoption_q1 = db.session.query(func.avg(Data.ADOPTION)).filter(Data.ADOPTION > 0).filter(Data.QUARTER == "Q12020").scalar()
        average_adoption_q2 = db.session.query(func.avg(Data.ADOPTION)).filter(Data.ADOPTION > 0).filter(Data.QUARTER == "Q22020").scalar()
        average_adoption_q3 = db.session.query(func.avg(Data.ADOPTION)).filter(Data.ADOPTION > 0).filter(Data.QUARTER == "Q32020").scalar()
        average_adoption_q4 = db.session.query(func.avg(Data.ADOPTION)).filter(Data.ADOPTION > 0).filter(Data.QUARTER == "Q42020").scalar()
        average_adoption_q5 = db.session.query(func.avg(Data.ADOPTION)).filter(Data.ADOPTION > 0).filter(Data.QUARTER == "Q12021").scalar()

        # print(" Avg adoption Q1 : %s ", average_adoption_q1)
        # print(" Avg adoption Q2 : %s ", average_adoption_q2)
        # print(" Avg adoption Q3 : %s ", average_adoption_q3)
        # print(" Avg adoption Q4 : %s ", average_adoption_q4)
        # print(" Avg adoption Q5 : %s ", average_adoption_q5)

        Q1Data = {}
        Q1Data["BPs"] = num_of_active_bps_q1
        Q1Data["APPs"] = num_of_running_APPs_q1
        Q1Data["ActiveVMs"] = num_of_active_AHV_VMs_q1
        Q1Data["name"] = "Q12020"

        Q2Data = {}
        Q2Data["BPs"] = num_of_active_bps_q2
        Q2Data["APPs"] = num_of_running_APPs_q2
        Q2Data["ActiveVMs"] = num_of_active_AHV_VMs_q2
        Q2Data["name"] = "Q22020"

        Q3Data = {}
        Q3Data["BPs"] = num_of_active_bps_q3
        Q3Data["APPs"] = num_of_running_APPs_q3
        Q3Data["ActiveVMs"] = num_of_active_AHV_VMs_q3
        Q3Data["name"] = "Q32020"

        Q4Data = {}
        Q4Data["BPs"] = num_of_active_bps_q4
        Q4Data["APPs"] = num_of_running_APPs_q4
        Q4Data["ActiveVMs"] = num_of_active_AHV_VMs_q4
        Q4Data["name"] = "Q42020"

        Q5Data = {}
        Q5Data["BPs"] = num_of_active_bps_q5
        Q5Data["APPs"] = num_of_running_APPs_q5
        Q5Data["ActiveVMs"] = num_of_active_AHV_VMs_q5
        Q5Data["name"] = "Q12021"

        statsByQtr.append(Q1Data)
        statsByQtr.append(Q2Data)
        statsByQtr.append(Q3Data)
        statsByQtr.append(Q4Data)
        statsByQtr.append(Q5Data)

        # print(statsByQtr)
    
    return jsonify({'statsByQtr': statsByQtr})

@app.route("/getProviderStatsByQtr", methods=['GET']) 
def getProviderStatsByQtr():
    providerStatsByQtr = []
    if request.method=='GET':
        num_of_active_AHV_VMs_q1 = db.session.query(func.sum(Data.ACTIVE_AHV_VMS)).filter(Data.QUARTER == "Q12020").scalar()
        # print(" Total Active BPs in Q1: %s " % num_of_active_AHV_VMs_q1)

        num_of_active_VMWare_VMs_q1 = db.session.query(func.sum(Data.ACTIVE_VMWARE_VMS)).filter(Data.QUARTER == "Q12020").scalar()
        # print(" Total Active BPs in Q1: %s " % num_of_active_VMWare_VMs_q1)

        num_of_active_AWS_VMs_q1 = db.session.query(func.sum(Data.ACTIVE_AWS_VMS)).filter(Data.QUARTER == "Q12020").scalar()
        # print(" Total Active BPs in Q1: %s " % num_of_active_AWS_VMs_q1)

        num_of_active_AZURE_VMs_q1 = db.session.query(func.sum(Data.ACTIVE_AZURE_VMS)).filter(Data.QUARTER == "Q12020").scalar()
        # print(" Total Active BPs in Q1: %s " % num_of_active_AZURE_VMs_q1)

        num_of_active_GCP_VMs_q1 = db.session.query(func.sum(Data.ACTIVE_GCP_VMS)).filter(Data.QUARTER == "Q12020").scalar()
        # print(" Total Active BPs in Q1: %s " % num_of_active_GCP_VMs_q1)

        num_of_active_AHV_VMs_q2 = db.session.query(func.sum(Data.ACTIVE_AHV_VMS)).filter(Data.QUARTER == "Q22020").scalar()
        # print(" Total Active BPs in Q2: %s " % num_of_active_AHV_VMs_q2)

        num_of_active_VMWare_VMs_q2 = db.session.query(func.sum(Data.ACTIVE_VMWARE_VMS)).filter(Data.QUARTER == "Q22020").scalar()
        # print(" Total Active BPs in Q2: %s " % num_of_active_VMWare_VMs_q2)

        num_of_active_AWS_VMs_q2 = db.session.query(func.sum(Data.ACTIVE_AWS_VMS)).filter(Data.QUARTER == "Q22020").scalar()
        # print(" Total Active BPs in Q2: %s " % num_of_active_AWS_VMs_q2)

        num_of_active_AZURE_VMs_q2 = db.session.query(func.sum(Data.ACTIVE_AZURE_VMS)).filter(Data.QUARTER == "Q22020").scalar()
        # print(" Total Active BPs in Q2: %s " % num_of_active_AZURE_VMs_q2)

        num_of_active_GCP_VMs_q2 = db.session.query(func.sum(Data.ACTIVE_GCP_VMS)).filter(Data.QUARTER == "Q22020").scalar()
        # print(" Total Active BPs in Q2: %s " % num_of_active_GCP_VMs_q2)

        num_of_active_AHV_VMs_q3 = db.session.query(func.sum(Data.ACTIVE_AHV_VMS)).filter(Data.QUARTER == "Q32020").scalar()
        # print(" Total Active BPs in Q3: %s " % num_of_active_AHV_VMs_q3)

        num_of_active_VMWare_VMs_q3 = db.session.query(func.sum(Data.ACTIVE_VMWARE_VMS)).filter(Data.QUARTER == "Q32020").scalar()
        # print(" Total Active BPs in Q3: %s " % num_of_active_VMWare_VMs_q3)

        num_of_active_AWS_VMs_q3 = db.session.query(func.sum(Data.ACTIVE_AWS_VMS)).filter(Data.QUARTER == "Q32020").scalar()
        # print(" Total Active BPs in Q3: %s " % num_of_active_AWS_VMs_q3)

        num_of_active_AZURE_VMs_q3 = db.session.query(func.sum(Data.ACTIVE_AZURE_VMS)).filter(Data.QUARTER == "Q32020").scalar()
        # print(" Total Active BPs in Q3: %s " % num_of_active_AZURE_VMs_q3)

        num_of_active_GCP_VMs_q3 = db.session.query(func.sum(Data.ACTIVE_GCP_VMS)).filter(Data.QUARTER == "Q32020").scalar()
        # print(" Total Active BPs in Q3: %s " % num_of_active_GCP_VMs_q3)

        num_of_active_AHV_VMs_q4 = db.session.query(func.sum(Data.ACTIVE_AHV_VMS)).filter(Data.QUARTER == "Q42020").scalar()
        # print(" Total Active BPs in Q4: %s " % num_of_active_AHV_VMs_q4)

        num_of_active_VMWare_VMs_q4 = db.session.query(func.sum(Data.ACTIVE_VMWARE_VMS)).filter(Data.QUARTER == "Q42020").scalar()
        # print(" Total Active BPs in Q4: %s " % num_of_active_VMWare_VMs_q4)

        num_of_active_AWS_VMs_q4 = db.session.query(func.sum(Data.ACTIVE_AWS_VMS)).filter(Data.QUARTER == "Q42020").scalar()
        # print(" Total Active BPs in Q4: %s " % num_of_active_AWS_VMs_q4)

        num_of_active_AZURE_VMs_q4 = db.session.query(func.sum(Data.ACTIVE_AZURE_VMS)).filter(Data.QUARTER == "Q42020").scalar()
        # print(" Total Active BPs in Q4: %s " % num_of_active_AZURE_VMs_q4)

        num_of_active_GCP_VMs_q4 = db.session.query(func.sum(Data.ACTIVE_GCP_VMS)).filter(Data.QUARTER == "Q42020").scalar()
        # print(" Total Active BPs in Q4: %s " % num_of_active_GCP_VMs_q4)

        num_of_active_AHV_VMs_q5 = db.session.query(func.sum(Data.ACTIVE_AHV_VMS)).filter(Data.QUARTER == "Q12021").scalar()
        # print(" Total Active BPs in Q4: %s " % num_of_active_AHV_VMs_q5)

        num_of_active_VMWare_VMs_q5 = db.session.query(func.sum(Data.ACTIVE_VMWARE_VMS)).filter(Data.QUARTER == "Q12021").scalar()
        # print(" Total Active BPs in Q4: %s " % num_of_active_VMWare_VMs_q5)

        num_of_active_AWS_VMs_q5 = db.session.query(func.sum(Data.ACTIVE_AWS_VMS)).filter(Data.QUARTER == "Q12021").scalar()
        # print(" Total Active BPs in Q4: %s " % num_of_active_AWS_VMs_q5)

        num_of_active_AZURE_VMs_q5 = db.session.query(func.sum(Data.ACTIVE_AZURE_VMS)).filter(Data.QUARTER == "Q12021").scalar()
        # print(" Total Active BPs in Q4: %s " % num_of_active_AZURE_VMs_q5)

        num_of_active_GCP_VMs_q5 = db.session.query(func.sum(Data.ACTIVE_GCP_VMS)).filter(Data.QUARTER == "Q12021").scalar()
        # print(" Total Active BPs in Q4: %s " % num_of_active_GCP_VMs_q5)

        Q1Data = {}
        Q1Data["AHV"] = num_of_active_AHV_VMs_q1
        Q1Data["VMWare"] = num_of_active_VMWare_VMs_q1
        Q1Data["AWS"] = num_of_active_AWS_VMs_q1
        Q1Data["AZURE"] = num_of_active_AZURE_VMs_q1
        Q1Data["GCP"] = num_of_active_GCP_VMs_q1
        Q1Data["name"] = "Q12020"

        Q2Data = {}
        Q2Data["AHV"] = num_of_active_AHV_VMs_q2
        Q2Data["VMWare"] = num_of_active_VMWare_VMs_q2
        Q2Data["AWS"] = num_of_active_AWS_VMs_q2
        Q2Data["AZURE"] = num_of_active_AZURE_VMs_q2
        Q2Data["GCP"] = num_of_active_GCP_VMs_q2
        Q2Data["name"] = "Q22020"

        Q3Data = {}
        Q3Data["AHV"] = num_of_active_AHV_VMs_q3
        Q3Data["VMWare"] = num_of_active_VMWare_VMs_q3
        Q3Data["AWS"] = num_of_active_AWS_VMs_q3
        Q3Data["AZURE"] = num_of_active_AZURE_VMs_q3
        Q3Data["GCP"] = num_of_active_GCP_VMs_q3
        Q3Data["name"] = "Q32020"

        Q4Data = {}
        Q4Data["AHV"] = num_of_active_AHV_VMs_q4
        Q4Data["VMWare"] = num_of_active_VMWare_VMs_q4
        Q4Data["AWS"] = num_of_active_AWS_VMs_q4
        Q4Data["AZURE"] = num_of_active_AZURE_VMs_q4
        Q4Data["GCP"] = num_of_active_GCP_VMs_q4
        Q4Data["name"] = "Q42020"

        Q5Data = {}
        Q5Data["AHV"] = num_of_active_AHV_VMs_q5
        Q5Data["VMWare"] = num_of_active_VMWare_VMs_q5
        Q5Data["AWS"] = num_of_active_AWS_VMs_q5
        Q5Data["AZURE"] = num_of_active_AZURE_VMs_q5
        Q5Data["GCP"] = num_of_active_GCP_VMs_q5
        Q5Data["name"] = "Q12021"

        providerStatsByQtr.append(Q1Data)
        providerStatsByQtr.append(Q2Data)
        providerStatsByQtr.append(Q3Data)
        providerStatsByQtr.append(Q4Data)
        providerStatsByQtr.append(Q5Data)

        # print(providerStatsByQtr)
    
    return jsonify({'finalrecord': providerStatsByQtr})

@app.route("/getSoftDeleteCustomers", methods=['GET']) 
def getSoftDeleteCustomers():
    soft_Delete_Customers = []
    if request.method=='GET':
        records = db.session.query(Data.Customer_Name, Data.PC_Cluster_UUID, Data.ACTIVE_AHV_VMS, Data.TOTAL_AHV_VMS, Data.RUNNING_APP, Data.PERCENT_VMs_INUSE).filter(Data.QUARTER == quarter_to_show).filter(Data.PERCENT_VMs_INUSE < 20).filter(Data.PERCENT_VMs_INUSE > -1).all()
        # print(" Soft Delete Customers are : %s " % records)

        for record in records:
            soft_delete_record = {}
            soft_delete_record["Customer"] = record.Customer_Name
            soft_delete_record["Cluster_ID"] = record.PC_Cluster_UUID
            soft_delete_record["Active_AHV_VMs"] = record.ACTIVE_AHV_VMS
            soft_delete_record["Total_AHV_VMs"] = record.TOTAL_AHV_VMS
            soft_delete_record["Running_App"] = record.RUNNING_APP
            soft_delete_record["Percent_InUse"] = record.PERCENT_VMs_INUSE
            
            soft_Delete_Customers.append(soft_delete_record)

        # print(soft_Delete_Customers)

    return jsonify({'Soft Delete Customers': soft_Delete_Customers})

@app.route("/getPublicAccounts", methods=['GET']) 
def getPublicAccounts():
    public_accounts_records = []
    if request.method=='GET':
        records = db.session.query(Data.Customer_Name, Data.AWS_ACCOUNT, Data.AZURE_ACCOUNT, Data.GCP_ACCOUNT, Data.PUBLIC_ACCOUNT).filter(Data.QUARTER == quarter_to_show).filter(Data.PUBLIC_ACCOUNT > -1).all()
        # print(" Public Account Customers are : %s " % records)

        for record in records:
            public_accounts_record = {}
            public_accounts_record["Customer"] = record.Customer_Name
            public_accounts_record["AWS"] = record.AWS_ACCOUNT
            public_accounts_record["AZURE"] = record.AZURE_ACCOUNT
            public_accounts_record["GCP"] = record.GCP_ACCOUNT
            public_accounts_record["PUBLIC_ACCOUNT"] = record.PUBLIC_ACCOUNT

            public_accounts_records.append(public_accounts_record)

        # print(public_accounts_records)
    
    return jsonify({'Public Account Records': public_accounts_records})

@app.route("/getCustomerDetails/<name>", methods=['GET']) 
def getCustomerDetails(name):
    custDetails = []
    if request.method=='GET':
        print("Customer name is : %s ", name);     
        records = db.session.query(Data.QUARTER, func.sum(Data.ACTIVE_BP).label("BPs"),func.sum(Data.RUNNING_APP).label("APPs"),func.sum(Data.LICENSE_UNIQUE_VMS_COUNT).label("VMs"),func.sum(Data.LICENSE_REQUIRED_PACKS).label("ReqLicense")).filter(Data.Customer_Name == name).group_by(Data.QUARTER).all()
        print(" Details for Customer : %s " % records)

        index = {'Q12020':0,'Q22020':1,'Q32020':2,'Q42020':3,'Q12021':4} 
        sortedRecords = [None] * len(index)
        print("Index size", len(index))
        print("SortedRecord size", len(sortedRecords))
        
        # sort records by quarter
        curIndex = 0
        for record in records:
            print("Record Quarter", record.QUARTER)
            sortedRecords[index.get(record.QUARTER)] = record; 
        
        for record in sortedRecords:
            custDetailRecord = {}
            if(record is not None):
                custDetailRecord["name"] = record.QUARTER
                custDetailRecord["BPs"] = record.BPs
                custDetailRecord["APPs"] = record.APPs
                custDetailRecord["VMs"] = record.VMs
        
            custDetails.append(custDetailRecord)

        print(custDetails)

    return jsonify({'Customer Details': custDetails})

@app.route("/getCustomerExtraDetails/<name>", methods=['GET']) 
def getCustomerExtraDetails(name):
    if request.method=='GET':
        # print("Customer name is : %s ", name);     
        records = db.session.query(Data.CALM_VERSION, Data.Last_Reported_Date, Data.ADOPTION).filter(Data.Customer_Name == name).filter(Data.QUARTER == quarter_to_show).all()
        
        for record in records:
            print ("Version", record.CALM_VERSION)
            version = record.CALM_VERSION
            print ("Date ", record.Last_Reported_Date)
            lastDate = record.Last_Reported_Date
            adoption = record.ADOPTION
        # print(" Details for Customer : %s " % records[0])

        licenses_purchased = db.session.query(func.sum(SalesData.QTY_SOLD)).filter(SalesData.CUSTOMER_NAME == name).scalar()
        # print(" Details for Customer : %s " % licenses_purchased)

        support_data = db.session.query(func.count(SupportData.CASE_NUM)).filter(SupportData.CUSTOMER_NAME == name).scalar()
        # print(" Details for Customer Support : %s " % support_data)

        custExtraDetails = {}
        custExtraDetails['Version'] = version
        custExtraDetails['ReportedDate'] = lastDate.strftime("%m/%d/%Y")
        custExtraDetails['Licenses'] = licenses_purchased
        custExtraDetails['SupportCases'] = support_data
        custExtraDetails['Adoption'] = adoption
        
    return custExtraDetails

@app.route("/getReportedDataSinceDays/<num>", methods=['GET'])
def getReportedDataSinceDays(num):
    current_time = datetime.datetime.utcnow()
    num_weeks_ago = current_time - datetime.timedelta(weeks=int(num))
    # within_Range = db.session.query(Data).filter(Data.QUARTER == quarter_to_show).filter(Data.Last_Reported_Date > num_weeks_ago).count()
    customer_Count = db.session.query(func.count(Data.Customer_Name)).filter(Data.QUARTER == quarter_to_show).scalar()
    within_Range = db.session.query(func.count(Data.Customer_Name)).filter(Data.QUARTER == quarter_to_show).scalar()
    
    # print(" Total Active Customers: %s " % customer_Count)
    # print(" Within Range Data: %s " % within_Range)
    # for record in within_Range:
    #     print("Customer is ", record.Customer_Name + "Date is ", record.Last_Reported_Date)

    sinceDays = {}
    sinceDays['Within Range'] = within_Range
    sinceDays['Total Customers'] = customer_Count

    return sinceDays

@app.route("/getLicenseData", methods=['GET'])
def getLicenseData():
    licenses_purchased = db.session.query(SalesData).all()
    license_records = []
    for record in licenses_purchased:
        license_record = {}
        license_record["CUSTOMER"] = record.CUSTOMER_NAME
        license_record["QTY_SOLD"] = record.QTY_SOLD
        license_record["QTR_SOLD"] = record.QTR_SOLD
        license_record["CALM_TCV"] = record.CALM_TCV
        license_record["SKU"] = record.PRODUCT_CODE
        

        license_records.append(license_record)

    # print(license_records)
    
    return jsonify({'License Records': license_records})

@app.route("/getTrialData", methods=['GET'])
def getTrialData():
    trial_customers = []
    
    licenses_purchased = db.session.query(SalesData.CUSTOMER_NAME).distinct().all()
    paid_customers = []
    for record in licenses_purchased:
        # print("Cust name ", record.CUSTOMER_NAME)
        if(record.CUSTOMER_NAME not in paid_customers):
            paid_customers.append(record.CUSTOMER_NAME)
   
    # print("Paid Customers ", paid_customers)
   
    pulse_records = db.session.query(Data.Customer_Name, Data.CALM_VERSION, Data.LICENSE_VMS_COUNTS, Data.Last_Reported_Date).filter(Data.QUARTER == quarter_to_show).all()
    for record in pulse_records:
        customer = record.Customer_Name
        if(customer not in paid_customers):
            trialrecord = {}
            trialrecord["CUSTOMER"] = record.Customer_Name
            trialrecord["Version"] = record.CALM_VERSION
            trialrecord["VMs"] = record.LICENSE_VMS_COUNTS
            tdate = record.Last_Reported_Date
            print ("TDate is ", tdate )
            trialrecord["Date"] = tdate.strftime("%m/%d/%Y")
        
            # trialrecord["Date"] = record.Last_Reported_Date
            trial_customers.append(trialrecord)
    print(trial_customers)
    return jsonify({'Trial Customers': trial_customers})

@app.route("/getSupportData", methods=['GET'])
def getSupportData():
    licenses_purchased = db.session.query(SalesData.CUSTOMER_NAME).distinct().all()
    paid_customers = []
    for record in licenses_purchased:
        # print("Cust name ", record.CUSTOMER_NAME)
        if(record.CUSTOMER_NAME not in paid_customers):
            paid_customers.append(record.CUSTOMER_NAME)
    
    # print("Paid Customers: ", paid_customers)

    support_data = db.session.query(SupportData).all()

    support_records = []
    support_dict = {}
    support_date_dict = {}
    for record in support_data:
        
        customer = record.CUSTOMER_NAME
        cDate = record.DATE
        # cDate = cDate.strftime("%Y/%m/%d")

        if(customer in support_dict):
            x = support_dict[customer]
            support_dict[customer] = x+1
        else:
            support_dict[customer] = 1
        
        if(customer in support_date_dict):
            # print(datetime.datetime.strptime(cDate, "%m/%d/%Y"))
            # print(datetime.datetime.strptime(support_date_dict[customer], "%m/%d/%Y"))
            # if((datetime.datetime.strptime(cDate), "%m/%d/%Y") > datetime.date(support_date_dict[customer])):
            if(cDate > support_date_dict[customer]):
                support_date_dict[customer] = cDate
        else:
            support_date_dict[customer] = cDate
        
    # print(support_dict)
    # return support_dict
    for key in support_dict:
        support_record = {}
        support_record["Customer_Name"] = key
        support_record["Value"] = support_dict[key]
        # tdate = datetime.datetime.strptime(support_date_dict[key].strftime('%Y-%m-%d'), '%Y-%d-%m')
        tdate = support_date_dict[key]
        print ("TDate is ", tdate )
        support_record["Date"] = tdate.strftime('%Y-%m-%d')
        # datetime.datetime.strftime(support_date_dict[key].strftime('%Y-%m-%d'), '%Y-%m-%d')
        if(key in paid_customers):
            support_record["Paid"] = "Yes"
        else:
            support_record["Paid"] = "No"

        support_records.append(support_record)

    # print(support_records)

    return jsonify({'Support records': support_records})

UPLOAD_FOLDER = '/Users/vishal.arhatia/pulse-transformer/'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route("/upload", methods=['POST'])
@cross_origin(origin='http://localhost:5000',headers=['Content- Type','Authorization'])
def fileUpload():
    print("I am in fileupload")
    # works through postman
    # file = request.files['file']
    # filename = file.filename
    # print("File name", filename)
    # destination=UPLOAD_FOLDER+filename
    # print(destination)
    # file.save(destination)

    file = request.files['file']
    print("File : ", file)
    print(type(file))
    filename = secure_filename(file.filename)
    # filename = file.filename
    print("File name is : ", filename)
    # print()
    
    response="Whatever you wish too return"
    
    return response

def process_new_SFDC_file():
    
    #Get Latest file
    list_of_files = glob.glob('*_sfdc.xlsx') 
    latest_file = max(list_of_files, key=os.path.getctime)
    print(latest_file)
    print(type(latest_file))

    db.session.query(SalesData).delete()
    db.session.commit()
    
    result = []
    for root, dir, files in os.walk('.'):
        if latest_file in files:
            result.append(os.path.join(root, latest_file))

    for excel in result:
        process_sfdc(excel)
    
    today = date.today()
    # Textual month, day and year	
    d2 = today.strftime("%B %d, %Y")
    print("d2 =", d2)
    print("d2 =", type(d2))

    global last_sfdc_update_time 
    last_sfdc_update_time = d2
    # total_tcv = db.session.query(func.sum(SalesData.CALM_TCV)).scalar()
    # print(total_tcv)

def process_new_PULSE_file():
    print("Processing new file")
    #Get Latest file
    list_of_files = glob.glob('*_pulse.xlsx') 
    latest_file = max(list_of_files, key=os.path.getctime)
    print(latest_file)
    print(type(latest_file))

    if(current_quarter == quarter_to_show):
        db.session.query(Data).filter(Data.QUARTER == quarter_to_show).delete()
        db.session.commit()
        
    result = []
    for root, dir, files in os.walk('.'):
        if latest_file in files:
            result.append(os.path.join(root, latest_file))

    for excel in result:
        process_pulse(excel)

    today = date.today()
    # Textual month, day and year	
    d2 = today.strftime("%B %d, %Y")
    
    global last_pulse_update_time
    last_pulse_update_time = d2

     
def process_new_SUPPORT_file():
    print("Processing new file")
    #Get Latest file
    list_of_files = glob.glob('*_cases.xlsx') 
    latest_file = max(list_of_files, key=os.path.getctime)
    print(latest_file)
    print(type(latest_file))

    db.session.query(SupportData).delete()
    db.session.commit()
        
    result = []
    for root, dir, files in os.walk('.'):
        if latest_file in files:
            result.append(os.path.join(root, latest_file))

    for excel in result:
        process_support(excel)

    today = date.today()
    # Textual month, day and year	
    d2 = today.strftime("%B %d, %Y")
    
    global last_support_update_time
    last_support_update_time = d2
 
@app.route("/getLastPulseUpdateTime", methods=['GET'])
def get_last_pulse_update_time():
    return jsonify({'Pulse Update Time': last_pulse_update_time})

@app.route("/getLastSFDCUpdateTime", methods=['GET'])
def get_last_sfdc_update_time():
    return jsonify({'SFDC Update Time': last_sfdc_update_time})

@app.route("/getLastSupportUpdateTime", methods=['GET'])
def get_last_support_update_time():
    return jsonify({'Support Update Time': last_support_update_time})

if __name__=="__main__":
    scheduler = BackgroundScheduler()
    # scheduler.add_job(func=process_newFile, trigger="interval", seconds=10)
    # scheduler.add_job(func=process_new_SFDC_file, trigger="interval", seconds=20)
    # scheduler.add_job(func=process_new_PULSE_file, trigger="interval", seconds=20)
    # scheduler.add_job(func=process_new_SUPPORT_file, trigger="interval", seconds=30)
    scheduler.start()

    app.run(debug=True, use_reloader=False)